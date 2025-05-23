from openpyxl import load_workbook, Workbook
import pdb

# Cargar los archivos
wb_original = load_workbook('entrega.xlsx')
ws_original = wb_original.active

wb_correcciones = load_workbook('corregidos.xlsx')
ws_correcciones = wb_correcciones.active

headers = [cell.value for cell in ws_original[1]]
rol_col_idx = headers.index('rol')
comuna_col_idx = headers.index('comuna')

correcciones = {}
rol_data = []
last_rol = ''
correcciones = {}
comuna_actual = None
rol_actual = None

for row in ws_correcciones.iter_rows(min_row=2, values_only=True):
    comuna, rol, predio, manzana, estado_o_cuota, monto = row
    
    if comuna and rol:
      # Es una nueva comuna/rol
      comuna_actual = comuna
      rol_actual = rol
      clave = (comuna_actual, rol_actual)
      correcciones['{}-{}-{}'.format(comuna_actual.lower(), int(rol_actual.split('-')[0]), int(rol_actual.split('-')[1]))] = [row]
    else:
       if comuna_actual and rol_actual:
          correcciones['{}-{}-{}'.format(comuna_actual.lower(), int(rol_actual.split('-')[0]), int(rol_actual.split('-')[1]))].append(row)

# Crear nuevo archivo Excel para los datos corregidos
wb_nuevo = Workbook()
ws_nuevo = wb_nuevo.active

# Escribir encabezado
ws_nuevo.append(headers)

for row in ws_original.iter_rows(min_row=2, values_only=True):
  rol = row[rol_col_idx]
  comuna = row[comuna_col_idx]
  if rol and '{}-{}-{}'.format(comuna.lower(), int(rol.split('-')[0]), int(rol.split('-')[1])) in correcciones:
     valores_corregidos = correcciones['{}-{}-{}'.format(comuna.lower(), int(rol.split('-')[0]), int(rol.split('-')[1]))]
     for fila in valores_corregidos:
        ws_nuevo.append(fila)
  else:
     ws_nuevo.append(row)
     
     
# Obtener todas las filas ya escritas
todas_las_filas = list(ws_nuevo.iter_rows(values_only=True)) 
# Procesar para eliminar líneas vacías consecutivas
filas_limpias = []
ultima_fue_vacia = False

for fila in todas_las_filas:
    es_vacia = all(campo is None or str(campo).strip() == '' for campo in fila)
    if es_vacia:
        if ultima_fue_vacia:
            continue
        else:
            filas_limpias.append(fila)
            ultima_fue_vacia = True
    else:
        filas_limpias.append(fila)
        ultima_fue_vacia = False

# Limpiar completamente la hoja actual
ws_nuevo.delete_rows(1, ws_nuevo.max_row)

# Reescribir las filas filtradas
for fila in filas_limpias:
    ws_nuevo.append(fila)

# Guardar
wb_nuevo.save('datos_corregidos.xlsx')
